"""
Генератор PDF исполнительных листов вертикальности опор.

Каждый PDF — один лист A3 (420×297 мм) для одной опоры.
Содержит:
- Мини-схема отклонения с красной линией
- Таблица координат проект/факт
- Штамп ПКФ СНАРК
- Заголовок с типом опоры

Также генерирует сводный Excel и ZIP-архив.
"""
from __future__ import annotations

import io
import math
import zipfile
from datetime import datetime
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle

from config import AppConfig, ASSETS_DIR


# ---------------------------------------------------------------------------
# Регистрация шрифтов (с fallback на стандартные)
# ---------------------------------------------------------------------------
_FONT_REGISTERED = False


def _register_fonts():
    global _FONT_REGISTERED
    if _FONT_REGISTERED:
        return
    try:
        pdfmetrics.registerFont(TTFont("DejaVuSans", "DejaVuSans.ttf"))
        pdfmetrics.registerFont(TTFont("DejaVuSans-Bold", "DejaVuSans-Bold.ttf"))
        _FONT_REGISTERED = True
    except Exception:
        _FONT_REGISTERED = True  # используем Helvetica как fallback


def _font(bold: bool = False) -> str:
    """Возвращает имя зарегистрированного шрифта."""
    _register_fonts()
    try:
        pdfmetrics.getFont("DejaVuSans")
        return "DejaVuSans-Bold" if bold else "DejaVuSans"
    except KeyError:
        return "Helvetica-Bold" if bold else "Helvetica"


# ---------------------------------------------------------------------------
# Константы вёрстки (A3 альбомная)
# ---------------------------------------------------------------------------
PAGE_W, PAGE_H = landscape(A3)  # ~1190.55 x 841.89 pt
MARGIN = 15 * mm
INNER_LEFT = 25 * mm  # левое поле (для подшивки)

# Зоны листа
STAMP_H = 60 * mm
TITLE_H = 20 * mm
SCHEME_W = 180 * mm
TABLE_W = PAGE_W - INNER_LEFT - MARGIN - SCHEME_W - 10 * mm


# ---------------------------------------------------------------------------
# Генерация PDF для одной опоры
# ---------------------------------------------------------------------------
def generate_pole_pdf(
    result: dict[str, Any],
    pdata: dict[str, Any],
    cfg: AppConfig,
) -> bytes:
    """
    Генерирует PDF исполнительного листа для одной опоры.

    Returns:
        bytes — содержимое PDF-файла.
    """
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=landscape(A3))
    c.setTitle(f"ИС вертикальности опоры {result['pole_name']}")

    _draw_frame(c, cfg)
    _draw_title(c, result, pdata, cfg)
    _draw_deviation_scheme(c, result, cfg)
    _draw_coordinates_table(c, result, cfg)
    _draw_stamp(c, result, pdata, cfg)
    _draw_deviation_info(c, result, cfg)

    c.save()
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Рамка ГОСТ
# ---------------------------------------------------------------------------
def _draw_frame(c: canvas.Canvas, cfg: AppConfig):
    """Рисует рамку листа по ГОСТ."""
    c.setStrokeColor(colors.black)

    # Внешняя рамка
    c.setLineWidth(0.5)
    c.rect(5 * mm, 5 * mm, PAGE_W - 10 * mm, PAGE_H - 10 * mm)

    # Внутренняя рамка
    c.setLineWidth(1.5)
    c.rect(INNER_LEFT, MARGIN, PAGE_W - INNER_LEFT - MARGIN, PAGE_H - 2 * MARGIN)


# ---------------------------------------------------------------------------
# Заголовок
# ---------------------------------------------------------------------------
def _draw_title(c: canvas.Canvas, result: dict, pdata: dict, cfg: AppConfig):
    """Рисует заголовок листа."""
    top = PAGE_H - MARGIN
    x = INNER_LEFT + 10 * mm

    # Логотип (если есть)
    logo_path = ASSETS_DIR / "logo_snark.png"
    if logo_path.exists():
        try:
            c.drawImage(
                str(logo_path), INNER_LEFT + 5 * mm, top - 18 * mm,
                width=15 * mm, height=15 * mm, preserveAspectRatio=True,
            )
            x = INNER_LEFT + 25 * mm
        except Exception:
            pass

    c.setFont(_font(bold=True), 14)
    c.drawString(x, top - 12 * mm, "ИСПОЛНИТЕЛЬНАЯ СХЕМА ВЕРТИКАЛЬНОСТИ ОПОРЫ")

    c.setFont(_font(), 11)
    pole_type = result.get("pole_type", "")
    c.drawString(
        x, top - 20 * mm,
        f'Опора {pole_type} № {result["pole_name"]}',
    )

    c.setFont(_font(), 9)
    c.drawString(
        x, top - 27 * mm,
        f'Проект: {pdata.get("name", "")}',
    )

    # Линия под заголовком
    c.setLineWidth(0.5)
    c.line(INNER_LEFT, top - 30 * mm, PAGE_W - MARGIN, top - 30 * mm)


# ---------------------------------------------------------------------------
# Мини-схема отклонения
# ---------------------------------------------------------------------------
def _draw_deviation_scheme(c: canvas.Canvas, result: dict, cfg: AppConfig):
    """
    Рисует мини-схему отклонения опоры:
    - Прямоугольник опоры (серый)
    - Осевая линия (штриховая)
    - Красная линия отклонения со стрелкой
    - Подписи ΔX, ΔY, угол
    """
    # Область схемы
    scheme_x = INNER_LEFT + 15 * mm
    scheme_y = MARGIN + STAMP_H + 15 * mm
    scheme_w = SCHEME_W - 30 * mm
    scheme_h = PAGE_H - 2 * MARGIN - TITLE_H - STAMP_H - 40 * mm

    # Заголовок схемы
    c.setFont(_font(bold=True), 11)
    c.drawString(scheme_x, scheme_y + scheme_h + 5 * mm, "Схема отклонения")

    # Рамка области
    c.setStrokeColor(colors.Color(0.8, 0.8, 0.8))
    c.setLineWidth(0.3)
    c.rect(scheme_x, scheme_y, scheme_w, scheme_h)

    # Центр схемы
    cx = scheme_x + scheme_w / 2
    cy = scheme_y + scheme_h / 2

    # Прямоугольник опоры (вид сверху — круг)
    pole_radius = min(scheme_w, scheme_h) * 0.06
    c.setStrokeColor(colors.Color(0.3, 0.3, 0.3))
    c.setFillColor(colors.Color(0.92, 0.92, 0.92))
    c.setLineWidth(1.5)
    c.circle(cx, cy, pole_radius, fill=1)

    # Перекрестие осей (проектное положение)
    axis_len = min(scheme_w, scheme_h) * 0.35
    c.setStrokeColor(colors.Color(0.5, 0.5, 0.5))
    c.setLineWidth(0.5)
    c.setDash(3, 3)
    c.line(cx - axis_len, cy, cx + axis_len, cy)
    c.line(cx, cy - axis_len, cx, cy + axis_len)
    c.setDash()

    # Подписи осей
    c.setFont(_font(), 8)
    c.setFillColor(colors.Color(0.4, 0.4, 0.4))
    c.drawString(cx + axis_len + 2 * mm, cy - 1 * mm, "X")
    c.drawString(cx - 3 * mm, cy + axis_len + 2 * mm, "Y")

    # Красная линия отклонения
    dx_mm = result.get("dx_mm", 0)
    dy_mm = result.get("dy_mm", 0)
    deviation_mm = result.get("deviation_mm", 0)

    if deviation_mm > 0.1:
        # Масштабирование: максимальное отклонение занимает 80% от axis_len
        scale = (axis_len * 0.8) / max(abs(dx_mm), abs(dy_mm), 1)

        end_x = cx + dx_mm * scale
        end_y = cy + dy_mm * scale

        c.setStrokeColor(colors.red)
        c.setLineWidth(2.5)
        c.line(cx, cy, end_x, end_y)

        # Стрелка
        _draw_arrow(c, cx, cy, end_x, end_y, 4 * mm)

        # Красная точка (фактическое положение)
        c.setFillColor(colors.red)
        c.circle(end_x, end_y, 2.5, fill=1, stroke=0)

        # Подписи отклонений
        c.setFont(_font(bold=True), 10)
        c.setFillColor(colors.red)

        label_x = end_x + 5 * mm
        label_y = end_y + 3 * mm
        c.drawString(label_x, label_y, f"{deviation_mm:.1f} мм")

        c.setFont(_font(), 8)
        c.setFillColor(colors.Color(0.3, 0.3, 0.3))
        c.drawString(label_x, label_y - 4 * mm, f"ΔX={dx_mm:.1f}  ΔY={dy_mm:.1f}")
        c.drawString(label_x, label_y - 8 * mm, f"Угол: {result.get('angle_deg', 0):.1f}°")

    else:
        c.setFont(_font(), 10)
        c.setFillColor(colors.Color(0.2, 0.6, 0.2))
        c.drawString(cx - 20 * mm, cy - pole_radius - 10 * mm, "Отклонение: 0.0 мм")

    # Вид сбоку (силуэт опоры)
    _draw_pole_side_view(c, scheme_x + scheme_w - 45 * mm, scheme_y + 10 * mm, result)

    c.setFillColor(colors.black)


def _draw_arrow(c: canvas.Canvas, x1: float, y1: float, x2: float, y2: float, size: float):
    """Рисует стрелку на конце линии."""
    angle = math.atan2(y2 - y1, x2 - x1)
    c.setFillColor(colors.red)

    p = c.beginPath()
    p.moveTo(x2, y2)
    p.lineTo(
        x2 - size * math.cos(angle - 0.3),
        y2 - size * math.sin(angle - 0.3),
    )
    p.lineTo(
        x2 - size * math.cos(angle + 0.3),
        y2 - size * math.sin(angle + 0.3),
    )
    p.close()
    c.drawPath(p, fill=1, stroke=0)


def _draw_pole_side_view(c: canvas.Canvas, x: float, y: float, result: dict):
    """Рисует вид сбоку опоры с отклонением."""
    h = 80 * mm
    w = 6 * mm

    # Подпись
    c.setFont(_font(bold=True), 8)
    c.setFillColor(colors.Color(0.3, 0.3, 0.3))
    c.drawString(x - 5 * mm, y + h + 5 * mm, "Вид сбоку")

    # Основание
    c.setStrokeColor(colors.Color(0.4, 0.4, 0.4))
    c.setFillColor(colors.Color(0.85, 0.85, 0.85))
    c.setLineWidth(1)
    c.rect(x - w * 1.5, y, w * 3, 3 * mm, fill=1)

    # Осевая (вертикаль проектная)
    c.setStrokeColor(colors.Color(0.6, 0.6, 0.6))
    c.setLineWidth(0.5)
    c.setDash(4, 2)
    c.line(x, y, x, y + h)
    c.setDash()

    # Опора (наклонённая)
    deviation_mm = result.get("deviation_mm", 0)
    dx_mm = result.get("dx_mm", 0)

    max_offset = 15 * mm
    offset = min(abs(dx_mm) * 0.15, max_offset) * (1 if dx_mm >= 0 else -1)

    c.setStrokeColor(colors.Color(0.3, 0.3, 0.3))
    c.setFillColor(colors.Color(0.75, 0.78, 0.82))
    c.setLineWidth(1.5)

    p = c.beginPath()
    p.moveTo(x - w / 2, y + 3 * mm)
    p.lineTo(x + offset - w / 2, y + h)
    p.lineTo(x + offset + w / 2, y + h)
    p.lineTo(x + w / 2, y + 3 * mm)
    p.close()
    c.drawPath(p, fill=1, stroke=1)

    # Красная линия отклонения вверху
    if abs(deviation_mm) > 0.1:
        c.setStrokeColor(colors.red)
        c.setLineWidth(1.5)
        c.line(x, y + h, x + offset, y + h)

        # Подпись
        c.setFont(_font(bold=True), 8)
        c.setFillColor(colors.red)
        side = x + offset + 3 * mm if offset >= 0 else x + offset - 20 * mm
        c.drawString(side, y + h - 2 * mm, f"{deviation_mm:.1f}")

    # Подписи высоты
    c.setFont(_font(), 7)
    c.setFillColor(colors.Color(0.3, 0.3, 0.3))
    c.drawString(x + w + 2 * mm, y + h / 2, f"H={result.get('height_project', 0):.1f} м")

    c.setFillColor(colors.black)


# ---------------------------------------------------------------------------
# Таблица координат
# ---------------------------------------------------------------------------
def _draw_coordinates_table(c: canvas.Canvas, result: dict, cfg: AppConfig):
    """Рисует таблицу координат проект/факт."""
    table_x = INNER_LEFT + SCHEME_W + 5 * mm
    table_y = MARGIN + STAMP_H + 15 * mm
    available_w = PAGE_W - MARGIN - table_x

    c.setFont(_font(bold=True), 11)
    table_top = PAGE_H - 2 * MARGIN - TITLE_H - 35 * mm
    c.drawString(table_x, table_top + 8 * mm, "Результаты контроля вертикальности")

    # Данные таблицы
    data = [
        ["Параметр", "Значение"],
        ["№ опоры", str(result.get("pole_name", ""))],
        ["Тип опоры", str(result.get("pole_type", ""))],
        ["Высота проект (м)", f"{result.get('height_project', 0):.3f}"],
        ["Высота факт (м)", f"{result.get('height_fact', 0):.3f}"],
        ["", ""],
        ["X проект (м)", f"{result.get('x_project', 0):.3f}"],
        ["Y проект (м)", f"{result.get('y_project', 0):.3f}"],
        ["X факт низ (м)", f"{result.get('x_fact_low', 0):.3f}"],
        ["Y факт низ (м)", f"{result.get('y_fact_low', 0):.3f}"],
        ["X факт верх (м)", f"{result.get('x_fact_high', 0):.3f}"],
        ["Y факт верх (м)", f"{result.get('y_fact_high', 0):.3f}"],
        ["", ""],
        ["ΔX (мм)", f"{result.get('dx_mm', 0):.1f}"],
        ["ΔY (мм)", f"{result.get('dy_mm', 0):.1f}"],
        ["Отклонение (мм)", f"{result.get('deviation_mm', 0):.1f}"],
        ["Угол (°)", f"{result.get('angle_deg', 0):.1f}"],
        ["Допуск (мм)", f"{result.get('tolerance_mm', 0):.1f}"],
        ["Статус", str(result.get("status", ""))],
    ]

    col_widths = [available_w * 0.55, available_w * 0.45]
    t = Table(data, colWidths=col_widths)

    status = result.get("status", "")
    status_color = (
        colors.Color(0.85, 1, 0.85) if status == "Норма"
        else colors.Color(1, 0.95, 0.8) if status == "Предупреждение"
        else colors.Color(1, 0.85, 0.85)
    )

    style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.1, 0.2, 0.36)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), _font(bold=True)),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("FONTNAME", (0, 1), (-1, -1), _font()),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("ALIGN", (1, 0), (1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.Color(0.7, 0.7, 0.7)),
        ("ROWHEIGHT", (0, 0), (-1, -1), 5.5 * mm),
        ("BACKGROUND", (0, -1), (-1, -1), status_color),
        ("FONTNAME", (0, -1), (-1, -1), _font(bold=True)),
    ])
    t.setStyle(style)

    t.wrapOn(c, available_w, 300 * mm)
    t.drawOn(c, table_x, table_top - len(data) * 5.5 * mm)


# ---------------------------------------------------------------------------
# Информация об отклонении (текстовый блок)
# ---------------------------------------------------------------------------
def _draw_deviation_info(c: canvas.Canvas, result: dict, cfg: AppConfig):
    """Информационный блок: метод измерения, ГОСТ, примечания."""
    info_x = INNER_LEFT + SCHEME_W + 5 * mm
    info_y = MARGIN + STAMP_H + 15 * mm

    c.setFont(_font(bold=True), 9)
    c.drawString(info_x, info_y + 45 * mm, "Метод контроля:")

    c.setFont(_font(), 8)
    lines = [
        "Геодезический контроль вертикальности опоры",
        "выполнен в соответствии с ГОСТ Р 51872-2024.",
        f"Точек нижнего сечения: {result.get('n_lower', 0)}",
        f"Точек верхнего сечения: {result.get('n_upper', 0)}",
        f"Дата: {datetime.now().strftime('%d.%m.%Y')}",
    ]
    for i, line in enumerate(lines):
        c.drawString(info_x, info_y + 38 * mm - i * 4 * mm, line)


# ---------------------------------------------------------------------------
# Штамп ПКФ СНАРК
# ---------------------------------------------------------------------------
def _draw_stamp(c: canvas.Canvas, result: dict, pdata: dict, cfg: AppConfig):
    """Рисует штамп в нижней части листа."""
    sx = INNER_LEFT
    sy = MARGIN
    sw = PAGE_W - INNER_LEFT - MARGIN
    sh = STAMP_H

    # Рамка штампа
    c.setStrokeColor(colors.black)
    c.setLineWidth(1)
    c.rect(sx, sy, sw, sh)

    # Вертикальные линии (6 колонок)
    col_widths = [0.30, 0.15, 0.15, 0.12, 0.12, 0.16]
    x_pos = sx
    for cw in col_widths[:-1]:
        x_pos += sw * cw
        c.line(x_pos, sy, x_pos, sy + sh)

    # Горизонтальная линия посередине
    c.line(sx, sy + sh / 2, sx + sw, sy + sh / 2)

    # Заполнение штампа
    x0 = sx + 5 * mm
    stamp = cfg.stamp

    # Верхняя строка
    c.setFont(_font(bold=True), 12)
    c.drawString(x0, sy + sh - 12 * mm, stamp.organization)

    c.setFont(_font(), 9)
    c.drawString(x0, sy + sh - 20 * mm, stamp.document_type)

    c.setFont(_font(), 8)
    c.drawString(x0, sy + sh - 27 * mm, f"Опора {result.get('pole_type', '')} № {result['pole_name']}")

    # Колонки с подписями
    col_x = sx + sw * 0.30 + 3 * mm
    labels_top = ["Стадия", "Лист", "Листов", "Дата"]
    values_top = [stamp.stage, "1", "1", datetime.now().strftime("%d.%m.%Y")]

    for i, (label, value) in enumerate(zip(labels_top, values_top)):
        xx = col_x + i * sw * 0.12 + (i * 3 * mm)
        c.setFont(_font(), 7)
        c.drawString(xx, sy + sh - 8 * mm, label)
        c.setFont(_font(bold=True), 9)
        c.drawString(xx, sy + sh - 16 * mm, value)

    # Нижняя строка
    c.setFont(_font(), 7)
    labels_bottom = ["Разработал", "Проверил", "ГИП"]
    names_bottom = [stamp.surveyor or "________", stamp.checker or "________", stamp.chief_engineer or "________"]

    for i, (label, name) in enumerate(zip(labels_bottom, names_bottom)):
        xx = x0 + i * 60 * mm
        c.drawString(xx, sy + sh / 2 - 8 * mm, label)
        c.setFont(_font(), 9)
        c.drawString(xx, sy + sh / 2 - 16 * mm, name)
        c.setFont(_font(), 7)

    # ГОСТ
    c.setFont(_font(), 7)
    c.drawString(
        sx + sw - 80 * mm, sy + 3 * mm,
        "ГОСТ Р 51872-2024"
    )


# ---------------------------------------------------------------------------
# Сводный Excel
# ---------------------------------------------------------------------------
def generate_summary_excel(
    results: list[dict[str, Any]],
    pdata: dict[str, Any],
) -> bytes:
    """Генерирует сводную таблицу отклонений в формате Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Сводная таблица"

    # Заголовок
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = f"Сводная таблица отклонений вертикальности — {pdata.get('name', '')}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:J2")
    ws["A2"].value = f"Дата: {datetime.now().strftime('%d.%m.%Y')}  |  ГОСТ Р 51872-2024"
    ws["A2"].font = Font(size=10, italic=True)

    # Заголовки колонок
    headers = [
        "№ опоры", "Тип", "Высота проект (м)", "Высота факт (м)",
        "ΔX (мм)", "ΔY (мм)", "Отклонение (мм)", "Угол (°)",
        "Допуск (мм)", "Статус",
    ]
    header_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    # Данные
    status_fills = {
        "Норма": PatternFill(start_color="d4edda", fill_type="solid"),
        "Предупреждение": PatternFill(start_color="fff3cd", fill_type="solid"),
        "Превышение": PatternFill(start_color="f8d7da", fill_type="solid"),
    }

    for row_idx, r in enumerate(results, 5):
        values = [
            r.get("pole_name", ""),
            r.get("pole_type", ""),
            r.get("height_project", 0),
            r.get("height_fact", 0),
            r.get("dx_mm", 0),
            r.get("dy_mm", 0),
            r.get("deviation_mm", 0),
            r.get("angle_deg", 0),
            r.get("tolerance_mm", 0),
            r.get("status", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            if col == 10:  # Статус
                cell.fill = status_fills.get(str(val), PatternFill())
                cell.font = Font(bold=True)

    # Автоширина
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# ZIP-архив
# ---------------------------------------------------------------------------
def create_zip_archive(files: list[tuple[str, bytes]]) -> bytes:
    """
    Создаёт ZIP-архив из списка файлов.

    Args:
        files: список (имя_файла, содержимое_bytes)

    Returns:
        bytes ZIP-архива.
    """
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, content in files:
            zf.writestr(name, content)
    return buf.getvalue()
